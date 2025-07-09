import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Backfill Report Tool", layout="wide")
st.title("📄 Backfill Report Generator")

# Initialize session state
if "page" not in st.session_state:
    st.session_state.page = "zqm"

# ----------------- PAGE 1: Upload ZQM -----------------
zqm_raw_df = None
if st.session_state.page == "zqm":
    st.header("Step 1: Upload and Filter ZQM Job File")
    zqm_file = st.file_uploader("📁 Upload ZQM Job File", type=["xlsx"], key="zqm")

    if zqm_file:
        df = pd.read_excel(zqm_file)
        df.columns = df.columns.str.strip()
        zqm_raw_df = df.copy()

        # Store ZQM file for reuse
        st.session_state.zqm_df = df

        # Apply filters
        filtered_df = df[
            (df["GR Qty"] == 0) &
            (df["Status"].str.contains("rel", case=False, na=False)) &
            (~df["Status"].str.contains("teco", case=False, na=False))
        ]

        st.success(f"✅ Filtered {len(filtered_df)} rows")
        st.dataframe(filtered_df)

        # Create downloadable Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            filtered_df.to_excel(writer, index=False, sheet_name="Filtered")
        output.seek(0)

        st.download_button(
            label="🗕️ Download Filtered ZQM as Excel (paste into /n/scwm/mon material request items)",
            data=output,
            file_name="filtered_zqm.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if st.button("➡️ Proceed to Upload PMR & SOH Files"):
            st.session_state.page = "pmr_soh"
            st.rerun()

# ----------------- PAGE 2: Upload PMR and SOH -----------------
elif st.session_state.page == "pmr_soh":
    st.header("Step 2: Upload PMR and SOH Files")
    st.markdown(":arrow_forward: **Instructions:** First, get a PMR export from SAP. Then, get the SOH file containing stock data for all part numbers in the PMR.")

    pmr_file = st.file_uploader("📁 Upload PMR File", type=["xlsx"], key="pmr")
    soh_file = st.file_uploader("📁 Upload SOH File", type=["xlsx"], key="soh")

    if pmr_file and soh_file:
        pmr_df = pd.read_excel(pmr_file)
        soh_df = pd.read_excel(soh_file)

        # Drop specified columns from PMR
        columns_to_drop = [
            "Blocked (Overall)", "Stock Type", "Unit of Measure", "Document Number",
            "Operation or Activity", "Stor. Bin of Goods Mvt Posting", "Party Entitled to Dispose",
            "Production Supply Area", "Reservation Number", "Staging Method", "Requirement Start Date"
        ]
        pmr_df = pmr_df.drop(columns=[col for col in columns_to_drop if col in pmr_df.columns])

        # Add ZQM basic start/finish dates and serial number
        if "zqm_df" in st.session_state:
            zqm_df = st.session_state.zqm_df.copy()
            zqm_df.columns = zqm_df.columns.str.strip().str.lower()
            order_col = next((col for col in zqm_df.columns if "order" in col), None)
            start_col = next((col for col in zqm_df.columns if "start" in col), None)
            finish_col = next((col for col in zqm_df.columns if "finish" in col), None)
            serial_col = next((col for col in zqm_df.columns if "serial" in col), None)
            if order_col and start_col and finish_col:
                zqm_df[start_col] = pd.to_datetime(zqm_df[start_col]).dt.strftime("%m/%d/%Y")
                zqm_df[finish_col] = pd.to_datetime(zqm_df[finish_col]).dt.strftime("%m/%d/%Y")
                zqm_subset_cols = [order_col, start_col, finish_col]
                if serial_col:
                    zqm_subset_cols.append(serial_col)
                zqm_subset = zqm_df[zqm_subset_cols].drop_duplicates()
                new_col_names = ["Manufacturing Order", "Basic start date", "Basic finish date"]
                if serial_col:
                    new_col_names.append("Serial Number")
                zqm_subset.columns = new_col_names

                pmr_order_col = next((col for col in pmr_df.columns if "order" in col.lower()), "Manufacturing Order")
                zqm_subset.columns = [pmr_order_col] + new_col_names[1:]
                pmr_df = pmr_df.merge(zqm_subset, on=pmr_order_col, how="left")

        soh_df.columns = soh_df.columns.str.strip()
        valid_stock_types = ["F1", "F2", "F3", "F4", "Q3", "Q4"]
        valid_storage_types = [
            "900", "900K", "905", "909", "910", "912", "914", "919", "920",
            "927", "934", "940", "986", "987", "990", "990P", "991", "FGI",
            "GRB1", "PRO", "PTWY", "RI", "RMA"
        ]
        soh_df_filtered = soh_df[
            soh_df["Stock Type"].isin(valid_stock_types) &
            soh_df["Storage Type"].astype(str).isin(valid_storage_types)
        ]

        soh_pivot = pd.pivot_table(
            soh_df_filtered,
            index="Product",
            columns="Owner",
            values="Quantity",
            aggfunc="sum",
            fill_value=0
        ).reset_index()

        # Detect product column in PMR
        product_column = None
        if "Product" in pmr_df.columns:
            product_column = "Product"
        elif "Finished Product or Order Text" in pmr_df.columns:
            product_column = "Finished Product or Order Text"

        if product_column:
            product_sums = soh_pivot.set_index("Product")[[col for col in soh_pivot.columns if col in ["MR9191", "MR9192"]]]
            pmr_df["9191"] = pmr_df[product_column].map(product_sums["MR9191"] if "MR9191" in product_sums else 0).fillna(0)
            pmr_df["9192"] = pmr_df[product_column].map(product_sums["MR9192"] if "MR9192" in product_sums else 0).fillna(0)

        # Generate pivot tables
        pivot1 = pd.pivot_table(
            pmr_df,
            index="Manufacturing Order",
            columns="Staging Status",
            values=product_column,
            aggfunc="count",
            fill_value=0
        ).reset_index()
        pivot2 = pd.pivot_table(
            pmr_df,
            index="Manufacturing Order",
            columns="Goods Issue Status",
            values=product_column,
            aggfunc="count",
            fill_value=0
        ).reset_index()

        combined_df = pd.merge(pivot1, pivot2, on="Manufacturing Order", how="outer", suffixes=('_Staging', '_GI')).fillna(0)

        def classify_hit(row):
            staging_cols = [col for col in row.index if '_Staging' in col]
            gi_cols = [col for col in row.index if '_GI' in col]

            any_partial = any("Partially Completed" in col and row[col] > 0 for col in row.index)
            if any_partial:
                return "Pulled"

            only_completed_or_not_relevant = all(
                ("Completed" in col or "Not Relevant" in col) and row[col] > 0 or row[col] == 0
                for col in staging_cols + gi_cols
            ) and (row.get('Completed_Staging', 0) > 0 or row.get('Not Relevant_Staging', 0) > 0) and (row.get('Completed_GI', 0) > 0 or row.get('Not Relevant_GI', 0) > 0)

            only_not_started = all(
                row.get(col, 0) == 0 for col in staging_cols + gi_cols
                if not col.endswith('Not Started_Staging') and not col.endswith('Not Started_GI')
            ) and (row.get('Not Started_Staging', 0) > 0 or row.get('Not Started_GI', 0) > 0)

            if only_completed_or_not_relevant:
                return "Completed"
            elif only_not_started:
                return "Not Pulled"
            else:
                return "Pulled"

        combined_df['Hit'] = combined_df.apply(classify_hit, axis=1)

        pmr_df = pmr_df.merge(combined_df[['Manufacturing Order', 'Hit']], on="Manufacturing Order", how="left")
        hit_col = pmr_df.pop("Hit")
        pmr_df.insert(0, "Hit", hit_col)

        st.success("✅ SOH processed and merged with PMR successfully!")
        st.dataframe(pmr_df)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pmr_df.to_excel(writer, index=False, sheet_name="MASTER")
            soh_df_filtered.to_excel(writer, index=False, sheet_name="SOH Raw")
            soh_pivot.to_excel(writer, index=False, sheet_name="SOH Pivot")
            pivot1.to_excel(writer, startrow=0, startcol=0, index=False, sheet_name="Pivot Summary")
            pivot2.to_excel(writer, startrow=0, startcol=len(pivot1.columns) + 2, index=False, sheet_name="Pivot Summary")
            combined_df.to_excel(writer, index=False, sheet_name="Combined Pivot")

            # Conditional formatting for Hit column
            wb = writer.book
            ws = wb["MASTER"]
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            for col in ws.iter_cols(min_row=1, max_row=1):
                if col[0].value == "Hit":
                    hit_col_letter = col[0].column_letter
                if col[0].value == "9191":
                    col_9191 = col[0].column_letter
                if col[0].value == "9192":
                    col_9192 = col[0].column_letter

            for row in range(2, ws.max_row + 1):
                hit_cell = ws[f"{hit_col_letter}{row}"]
                soh_9191_cell = ws[f"{col_9191}{row}"]
                soh_9192_cell = ws[f"{col_9192}{row}"]

                if hit_cell.value == "Completed":
                    hit_cell.fill = green_fill
                elif hit_cell.value == "Pulled":
                    hit_cell.fill = red_fill
                elif hit_cell.value == "Not Pulled":
                    hit_cell.fill = yellow_fill

                if isinstance(soh_9191_cell.value, (int, float)) and soh_9191_cell.value > 0:
                    soh_9191_cell.fill = green_fill
                if isinstance(soh_9192_cell.value, (int, float)) and soh_9192_cell.value > 0:
                    soh_9192_cell.fill = green_fill

            writer.book["MASTER"].sheet_properties.tabColor = "00FF00"

        output.seek(0)

        st.download_button(
            label="📅 Download Updated PMR File",
            data=output,
            file_name="updated_pmr_with_soh.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if st.button("🔁 Restart Entire Process"):
            st.session_state.page = "zqm"
            st.rerun()